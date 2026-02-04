from django.shortcuts import render, redirect
# from blog.models import Blog  # Removed Django model usage

from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from utils.image_processor import ImageProcessor
from utils.dynamodb import (fetch_all_inquiries, fetch_all_appointments, update_item_status, 
                            delete_dynamo_item, get_dynamodb_resource, fetch_blogs, fetch_blog_by_id, 
                            save_blog, delete_blog, fetch_gallery_photos, save_gallery_photo, 
                            delete_gallery_photo, fetch_gallery_videos, save_gallery_video, 
                            delete_gallery_video)
import openpyxl
from openpyxl.utils import get_column_letter
import datetime
import uuid

@require_http_methods(['GET'])
def admin_dashboard_view(request):
    # Daily counters (simplified: scan and count for today)
    today_str = datetime.datetime.now().strftime('%Y-%m-%d')
    all_inquiries = fetch_all_inquiries()
    all_appointments = fetch_all_appointments()
    
    today_inquiries = [i for i in all_inquiries if i.get('created_at', '').startswith(today_str)]
    today_appointments = [a for a in all_appointments if a.get('date') == today_str]
    
    context = {
        'total_inquiries': len(all_inquiries),
        'total_appointments': len(all_appointments),
        'total_items': len(all_inquiries) + len(all_appointments),
        'today_inquiries_count': len(today_inquiries),
        'today_appointments_count': len(today_appointments),
    }
    return render(request, 'html/admin/dashboard.html', context)

@require_http_methods(['GET', 'POST'])
def admin_enquiry_list_view(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        items = request.POST.getlist('items[]')
        
        if action == 'delete':
            for item_id in items:
                delete_dynamo_item('Patients', item_id)
            return JsonResponse({'success': True})
            
        elif action == 'resolve':
            for item_id in items:
                update_item_status('Patients', item_id, 'resolved')
            return JsonResponse({'success': True})

    inquiries = fetch_all_inquiries()
    
    # Filter
    status_filter = request.GET.get('status')
    if status_filter:
        inquiries = [i for i in inquiries if i.get('status') == status_filter]
        
    return render(request, 'html/admin/enquiry_list.html', {'inquiries': inquiries})

@require_http_methods(['GET', 'POST'])
def admin_appointment_list_view(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        appointment_id = request.POST.get('appointment_id')
        notes = request.POST.get('notes', '')
        
        if action == 'confirm':
            update_item_status('Appointments', appointment_id, 'confirmed', notes)
            return JsonResponse({'success': True})
        elif action == 'cancel':
            update_item_status('Appointments', appointment_id, 'cancelled', notes)
            return JsonResponse({'success': True})
        elif action == 'delete':
            items = request.POST.getlist('items[]')
            for item_id in items:
                delete_dynamo_item('Appointments', item_id)
            return JsonResponse({'success': True})

    appointments = fetch_all_appointments()
    
    # Filter by Date (Today)
    date_filter = request.GET.get('date')
    if date_filter == 'today':
        today_str = datetime.datetime.now().strftime('%Y-%m-%d')
        appointments = [a for a in appointments if a.get('date') == today_str]
    
    # Filter by Status
    status_filter = request.GET.get('status')
    if status_filter:
        appointments = [a for a in appointments if a.get('status') == status_filter]

    return render(request, 'html/admin/appointment_list.html', {'appointments': appointments})

@require_http_methods(['GET'])
def export_excel_view(request, type):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    if type == 'enquiries':
        ws.title = "Enquiries"
        headers = ['ID', 'Name', 'Email', 'Phone', 'Description', 'Status', 'Date']
        data = fetch_all_inquiries()
        fields = ['id', 'name', 'email', 'phone', 'description', 'status', 'created_at']
    else:
        ws.title = "Appointments"
        headers = ['ID', 'Name', 'Email', 'Phone', 'Service', 'Date', 'Time', 'Status', 'Notes']
        data = fetch_all_appointments()
        fields = ['id', 'name', 'email', 'phone', 'service', 'date', 'time_slot', 'status', 'admin_notes']

    # Write headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Write data
    for row, item in enumerate(data, 2):
        for col, field in enumerate(fields, 1):
            ws.cell(row=row, column=col, value=str(item.get(field, '')))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={type}_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


# --- Blog Management ---

@require_http_methods(['GET', 'POST'])
def admin_blog_list_view(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        blog_id = request.POST.get('blog_id')
        
        if action == 'delete':
            delete_blog(blog_id)
            return JsonResponse({'success': True})
        
        elif action == 'toggle_draft':
            blog_item = fetch_blog_by_id(blog_id)
            if blog_item:
                blog_item['draft'] = not blog_item.get('draft', True)
                save_blog(blog_item)
                return JsonResponse({'success': True, 'new_status': blog_item['draft']})


    blogs = fetch_blogs(include_drafts=True)
    return render(request, 'html/admin/blog_list.html', {'blogs': blogs})


@require_http_methods(['GET', 'POST'])
def admin_blog_edit_view(request, blog_id=None):
    blog_obj = None
    if blog_id and blog_id != 'new':
        blog_obj = fetch_blog_by_id(blog_id)

    if request.method == 'POST':
        data = {
            'id': blog_id if blog_id != 'new' else str(uuid.uuid4()),
            'title': request.POST.get('title'),
            'slug': request.POST.get('slug'),
            'body': request.POST.get('body'),
            'thumbnail': request.POST.get('thumbnail'),
            'draft': request.POST.get('draft') == 'on',
            'meta_description': request.POST.get('meta_description'),
            'category': request.POST.get('category', 'General'),
        }
        save_blog(data)
        return redirect('admin_blogs')


    return render(request, 'html/admin/blog_edit.html', {'blog': blog_obj})


# --- Gallery Management ---

@require_http_methods(['GET', 'POST'])
def admin_gallery_photos_view(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            # Handle file upload or URL
            image_url = request.POST.get('image_url', '').strip()
            uploaded_file = request.FILES.get('image_file')
            
            if uploaded_file:
                # Process and save uploaded file
                try:
                    image_url = ImageProcessor.save_uploaded_image(uploaded_file, 'photos', 'photo')
                except ValueError as e:
                    # Handle validation error - for now just skip
                    pass
            
            if image_url:
                save_gallery_photo(
                    image_url,
                    request.POST.get('title', ''),
                    request.POST.get('description', '')
                )
        elif action == 'delete':
            delete_gallery_photo(request.POST.get('item_id'))
        return redirect('admin_gallery_photos')

    photos = fetch_gallery_photos()
    return render(request, 'html/admin/gallery_photos_manage.html', {'photos': photos})


@require_http_methods(['GET', 'POST'])
def admin_gallery_videos_view(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            from utils.youtube_utils import extract_youtube_id
            
            # Extract YouTube ID from URL
            video_input = request.POST.get('video_url', '').strip()
            video_id = extract_youtube_id(video_input)
            
            if not video_id:
                # Invalid URL/ID - strictly reject
                print("Invalid YouTube URL provided")
                return redirect('admin_gallery_videos')
            
            # Additional strict validation
            from utils.youtube_utils import validate_youtube_id
            if not validate_youtube_id(video_id):
                print(f"Invalid Video ID extracted: {video_id}")
                return redirect('admin_gallery_videos')
            
            # Save with auto-generated thumbnail
            save_gallery_video(
                video_id,
                request.POST.get('title', ''),
                request.POST.get('description', ''),
                ''  # Empty thumbnail_url will trigger auto-generation
            )
        elif action == 'delete':
            delete_gallery_video(request.POST.get('item_id'))
        return redirect('admin_gallery_videos')

    videos = fetch_gallery_videos()
    return render(request, 'html/admin/gallery_videos_manage.html', {'videos': videos})
